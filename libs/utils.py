from collections.abc import Iterable
from itertools import groupby
from numpy import ndarray, nan
from pandas import isnull
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.cell import get_column_letter, range_boundaries
from os import path, makedirs
from six import string_types


def deduplicate_list(src_list):
    """
    Gets rid of duplicates from a given list and preserves the order.

    :param src_list: source list
    :return: unique list
    """
    new_list = []
    for el in src_list:
        if el not in new_list:
            new_list.append(el)
    return new_list


def sort_dict(src_dict):
    """
    Sort given dictionary

    :param src_dict: source dict
    :return: sorted dictionary
    """
    sorted_dict = {k: src_dict[k] for k in sorted(src_dict.keys())}
    return sorted_dict


def join_strings(sequence, separator=',', omit_empty=True):
    """ Join sequence of strings into one string using specified separator. All non-string values in sequence are
    ignored.

    :param sequence: sequence of strings to join
    :param separator: separator used for joining. By default, "," (comma).
    :param omit_empty: ignore empty strings. By default, True.
    :return: joined string
    """
    sequence = [item for item in sequence if isinstance(item, str) and
                (item.strip() != '' and omit_empty or not omit_empty)]
    return separator.join(sequence)


def capitalize_string_words(src_string):
    """
    Capitalize all the words in the given string

    :param src_string: source string
    :return:
    """
    return ' '.join([s.capitalize() for s in src_string.split(' ')])


def get_dataframe_column_index(dataframe, column_name):
    """Get the list of column indexes matching specified column name.

    :param dataframe: source dataframe
    :type dataframe: pandas.DataFrame
    :param column_name: column name

    :return: matching column indexes
    :rtype: list of int
    """
    try:
        result = dataframe.columns.get_loc(column_name)
    except KeyError:
        return []
    if isinstance(result, int):
        return [result]
    # result contains slice
    elif isinstance(result, slice):
        # return sliced list of indexes
        return [i for i, _ in enumerate(dataframe.columns)][result]
    # result contains array of booleans
    elif isinstance(result, ndarray):
        # return list of column indexes where value is True
        return [i for i, b in enumerate(result) if b]
    else:
        return []


def _merge_cells_auto(worksheet, min_row=1, min_col=1, max_row=None, max_col=None):
    """Merge worksheet cells automatically.

    In case adjacent left-hand cell has the same value then current and adjacent cells get merged horizontally;
    vertically - if adjacent cell above have the same value, None or nan

    :param worksheet: target worksheet
    :type worksheet: openpyxl.worksheet.worksheet.Worksheet
    :param min_row: min row index of cells range for merging
    :param min_col: min column index
    :param max_row: max row index
    :param max_col: max column index
    :return: None
    """
    prev_hor_groups = None
    cells_to_merge = []
    # search for cells to merge horizontally
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col,
                                                      max_col=max_col)):
        # group top level columns
        if prev_hor_groups is None:
            group_func = lambda x: (None, x[1].value)
        # group columns respecting parent header bounds
        else:
            group_func = lambda x: (prev_hor_groups[x[0]], x[1].value)
        # create list of dict storing column name, row index, and tuple of column indexes to merge
        group = [{'val': key[1],
                  'row': row_idx,
                  'cols': tuple(g[0] for g in group)} for key, group in groupby(enumerate(row), key=group_func)]
        hor_groups = []
        # 'explode' grouped list into a list by the number of columns; new list contains dict storing column name,
        # and tuple of min and max column indexes
        for col_idx, col in enumerate(group):
            repeat = len(col['cols'])
            col['cols'] = (min(col['cols']), max(col['cols']))
            # replace None- and nan-values with value from the above level of header if lists of
            # column indexes are equal
            if col['val'] in (None, nan) and row_idx > 0:
                col['val'] = next((g['val'] for g in prev_hor_groups if g['cols'] == col['cols']), None)
            for _ in range(0, repeat):
                hor_groups.append(col)
        prev_hor_groups = hor_groups
        cells_to_merge.append(hor_groups)
    # search for cells to merge vertically
    for i in range(0, max_col - min_col + 1):
        col = [row[i] for row in cells_to_merge]
        # group by column values, min and max column indexes
        ver_group = [list(group) for key, group in groupby(col, key=lambda x: (x['val'], x['cols']))]
        for group in ver_group:
            # read min/max column indexes for merging from the very first element in group
            cols = group[0]['cols']
            # get list of row indexes for merging
            row_idxs = [r['row'] for r in group]
            rows = (min(row_idxs), max(row_idxs))
            # merge cells having at least 2 rows or columns to merge
            if rows[0] != rows[1] or cols[0] != cols[1]:
                worksheet.merge_cells(start_row=rows[0]+1, start_column=cols[0]+1, end_row=rows[1]+1,
                                      end_column=cols[1]+1)
                # if cells get merged horizontally then align merged cell centrally
                if cols[0] != cols[1]:
                    cell = worksheet.cell(row=rows[0]+1, column=cols[0]+1)
                    cell.alignment = Alignment(horizontal='center')


def write_dataframe_headers_to_excel(worksheet, dataframe, automerge=True, wrap_text=True):
    """Write DataFrame indexes names and column names to Excel worksheet.

    :param worksheet: target worksheet
    :type worksheet: openpyxl.worksheet.worksheet.Worksheet
    :param dataframe: source dataframe
    :type dataframe: pandas.DataFrame
    :param automerge: merge cells automatically
    :type automerge: boolean
    :param wrap_text: wrap text in cell
    :type wrap_text: boolean
    :return: None
    """
    # init styles for headers
    font_style = Font(bold=True)
    border_side = Side(border_style='thin', color='000000')
    border_style = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    alignment = Alignment(wrap_text=wrap_text)
    # determine how many levels do the columns contain
    header_levels = dataframe.columns.nlevels
    # fill up the indexes
    for idx, index_name in enumerate(dataframe.index.names):
        cell = worksheet.cell(row=1, column=idx + 1)
        cell.value = index_name
        cell.border = border_style
        cell.font = font_style
        cell.alignment = alignment
        if header_levels <= 1:
            continue
        # worksheet.merge_cells(start_row=1, start_column=idx + 1, end_row=header_levels, end_column=idx + 1)
    # fill up the columns
    col_offset = len(dataframe.index.names)
    col_count = len(dataframe.columns)
    for row in range(0, header_levels):
        for col in range(0, col_count):
            if header_levels == 1:
                col_name = dataframe.columns[col]
            else:
                col_name = dataframe.columns[col][row]
            cell = worksheet.cell(row=row+1, column=col+col_offset+1)
            cell.value = col_name
            cell.border = border_style
            cell.font = font_style
            cell.alignment = alignment
    if automerge:
        _merge_cells_auto(worksheet=worksheet, min_row=1, max_row=header_levels, min_col=1,
                          max_col=col_offset+col_count)
    # freeze header rows
    worksheet.freeze_panes = f'A{header_levels + 1}'
    # set autofilter on the bottom header row
    worksheet.auto_filter.ref = f'A{header_levels}:' \
                                f'{get_column_letter(col_offset + col_count)}{header_levels}'


def write_dataframe_values_to_excel(worksheet, dataframe, start_row=None, wrap_text=True):
    """Write DataFrame values to Excel worksheet

    :param worksheet: target Excel worksheet
    :type worksheet: openpyxl.worksheet.worksheet.Worksheet
    :param dataframe: source dataframe
    :type dataframe: pandas.DataFrame
    :param start_row: start row index for writing data
    :param wrap_text: wrap text in cell
    :type wrap_text: boolean
    :return: None
    """
    # init styles for values
    border_side = Side(border_style='thin', color='000000')
    border_style = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    alignment = Alignment(wrap_text=wrap_text)
    for row_idx, row_data in enumerate(dataframe.itertuples()):
        values_list = []
        row_cells = []
        for val in row_data:
            # value in dataframe may be a tuple in case of existing multiindex and it cannot be correctly written to
            # excel worksheet; so convert the tuple into a list and merge it with values list
            if isinstance(val, tuple):
                values_list += list(val)
            else:
                values_list.append(val)
        for col_idx, val in enumerate(values_list):
            cell = WriteOnlyCell(worksheet)
            cell.border = border_style
            if isinstance(val, Iterable) and not isinstance(val, string_types):
                cell.value = ', '.join([str(v) for v in val])
            else:
                cell.value = None if isnull(val) else val
            cell.alignment = alignment
            row_cells.append(cell)
        if start_row is None:
            worksheet.append(row_cells)
        else:
            worksheet.insert_rows(start_row, amount=1)
            for col_idx, cell in enumerate(row_cells):
                new_cell = worksheet.cell(row=start_row, column=col_idx+1)
                new_cell.value = cell.value
                new_cell.border = border_style
                new_cell.alignment = alignment
            start_row += 1


def set_excel_col_autowidth(worksheet, min_width=10, max_width=50):
    """Adjust column width automatically in a given Excel worksheet

    :param worksheet: target worksheet
    :type worksheet: openpyxl.worksheet.worksheet.Worksheet
    :param min_width: minimum width of columns
    :param max_width: maximum width of columns
    :return: None
    """
    merged_cells_bounds = [range_boundaries(str(c)) for c in worksheet.merged_cells.ranges if c.min_col != c.max_col]
    for col_idx, col in enumerate(worksheet.columns):
        col_idx += 1
        col_letter = get_column_letter(col_idx)
        current_width = min_width
        worksheet.column_dimensions[col_letter].width = min_width
        for row_idx, row in enumerate(col):
            row_idx += 1
            col_distr = next((b[2] - b[0] + 1 for b in merged_cells_bounds
                              if row_idx in range(b[1], b[3]+1) and col_idx in range(b[0], b[2]+1)), 1)
            cell = worksheet.cell(row=row_idx, column=col_idx)
            chars = len(str(cell.value))/col_distr if cell.value is not None else 0
            if cell.alignment.wrap_text:
                # if text is wrapped then get amount of chars from the longest line
                chars = max([len(l)/col_distr for l in str(cell.value).split('\n')])
            # 2.5 - approximate width of filter icon
            desired_width = min(chars * 1.1 + 2.5, max_width)
            if desired_width > current_width:
                current_width = desired_width
                worksheet.column_dimensions[col_letter].width = current_width


def save_dataframes_to_excel(filepath, sheets_dataframes, wb_append=False, wrap_text=True, omit_index=False):
    """Save dataframes to Excel workbook

    :param filepath: target path of Excel workbook
    :param sheets_dataframes: dataframes to save
    :type sheets_dataframes: dict, where key = sheet name, value = dataframe
    :param wb_append: append data to workbook if it already exists, otherwise - overwrite it
    :type wb_append: boolean
    :param wrap_text: wrap text in cell
    :type wrap_text: boolean
    :param omit_index: do not output dataframe index
    :type omit_index: boolean
    :return: None
    """
    dirpath = path.dirname(filepath)
    if not path.exists(dirpath):
        makedirs(dirpath, exist_ok=True)
    new_wb = True
    if path.exists(filepath) and wb_append:
        wb = load_workbook(filepath)
        new_wb = False
    else:
        wb = Workbook()
    for i, v in enumerate(sheets_dataframes.items()):
        sheet_name = v[0]
        df = v[1]
        # rename first worksheet if workbook should be overwritten
        if i == 0 and new_wb:
            ws = wb.active
            ws.title = sheet_name
        else:
            sheet_index = None
            try:
                # delete existing worksheet with the same name
                sheet_index = wb.sheetnames.index(sheet_name)
                wb.remove_sheet(wb[sheet_name])
            except ValueError:
                pass
            ws = wb.create_sheet(sheet_name, sheet_index)
        write_dataframe_headers_to_excel(worksheet=ws, dataframe=df, wrap_text=wrap_text)
        write_dataframe_values_to_excel(worksheet=ws, dataframe=df, wrap_text=wrap_text)
        set_excel_col_autowidth(ws)
    wb.save(filename=filepath)
